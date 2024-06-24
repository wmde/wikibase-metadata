"""Compile Spreadsheet"""

import os
from openpyxl import Workbook, load_workbook

from src.project_constants import COMPLETED_FILENAME, URL_FILENAME
from src.url_sheet import create_url_sheet
from src.user_data.user_data_sheet import create_user_data_sheet


def main():
    """Compile Spreadsheet"""

    if os.path.exists(URL_FILENAME):
        wb = load_workbook(URL_FILENAME)
    else:
        wb = Workbook()
    create_url_sheet(wb)

    create_user_data_sheet(wb)

    entity_sheet = wb.create_sheet("entity")
    entity_sheet.title = "Entity Data"

    wb.save(COMPLETED_FILENAME)


if __name__ == "__main__":
    main()
