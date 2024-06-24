from openpyxl import Workbook

from src.project_constants import USER_DATA_FILENAME
from src.user_data.compile_user_data import (
    compileAllImplicitUserGroups,
    compileAllUserGroups,
    compileUserGroupCounts,
)
from src.user_data.fetch_user_data import fetchUserData
from src.utils.assign_cell import assign_cell


def create_user_data_sheet(wb: Workbook):
    ws = wb.create_sheet("users")
    ws.title = "User Data"

    assign_cell(ws, 1, 1, "API URL")
    assign_cell(ws, 2, 1, "Data Returned")
    assign_cell(ws, 3, 1, "Total Count")
    assign_cell(ws, 4, 1, "Groups")
    assign_cell(ws, 5, 1, "Implicit Groups")

    wikibase_default_groups = [
        "*",
        "user",
        "autoconfirmed",
        "sysop",
        "bureaucrat",
        "interface-admin",
    ]
    for i, g in enumerate(wikibase_default_groups):
        assign_cell(ws, 6 + i, 1, g)

    url_ws = wb["URLS"]
    api_urls = [c.value for i, c in enumerate(url_ws["B:B"]) if i > 0]

    for i, api_url in enumerate(api_urls):
        assign_cell(ws, 1, i + 2, api_url)

        try:
            site_user_data = fetchUserData(api_url)
            assign_cell(ws, 2, i + 2, True)
        except:
            assign_cell(ws, 2, i + 2, False)
            print("\tFailed")
            continue

        assign_cell(ws, 3, i + 2, len(site_user_data))

        site_user_groups = compileAllUserGroups(site_user_data)
        print(f"\t{', '.join(site_user_groups)}")
        assign_cell(ws, 4, i + 2, ", ".join(site_user_groups))

        site_implicit_user_groups = compileAllImplicitUserGroups(site_user_data)
        print(f"\t{', '.join(site_implicit_user_groups)}")
        assign_cell(ws, 5, i + 2, ", ".join(site_implicit_user_groups))

        existing_spreadsheet_groups = {
            cell.value for i, cell in enumerate(ws["1:1"]) if i > 4
        }
        add_groups = site_user_groups - existing_spreadsheet_groups

        for j, g in enumerate(add_groups):
            print(
                f"\tAdding '{g}' in ({5 + len(existing_spreadsheet_groups) + j + 1}, {1})"
            )
            assign_cell(ws, 5 + len(existing_spreadsheet_groups) + j + 1, 1, g)

        updated_spreadsheet_groups = [
            cell.value for i, cell in enumerate(ws["1:1"]) if i > 4
        ]

        site_group_counts = compileUserGroupCounts(site_user_data)

        for k, v in site_group_counts.items():
            assign_cell(ws, 5 + updated_spreadsheet_groups.index(k) + 1, i + 2, v)

    wb.save(USER_DATA_FILENAME)
